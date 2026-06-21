import json
import os
import sqlite3
import sys
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "backend"))
os.environ.setdefault("WEBHOOK_HMAC_SECRET", "test-secret-32chars-aaaaaaaaaaaa")
os.environ.setdefault("DATABASE_URL", "sqlite:///./test_alerts.db")
os.environ.setdefault("TELEGRAM_BOT_TOKEN", "")
os.environ.setdefault("TELEGRAM_CHAT_ID", "")

from core.config import settings
from db.lifecycle import ro_connect
from main import app


SCHEMA = """
CREATE TABLE signal_lifecycle (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    setup TEXT NOT NULL, direction TEXT NOT NULL, pattern TEXT NOT NULL,
    killzone TEXT, killzone_match INTEGER NOT NULL,
    entry REAL NOT NULL, entry_fill REAL NOT NULL, spread_pips REAL NOT NULL DEFAULT 0,
    sl REAL NOT NULL, tp1 REAL NOT NULL, tp_final REAL NOT NULL,
    risk_pips REAL NOT NULL, planned_rr REAL NOT NULL,
    status TEXT NOT NULL, mfe_pips REAL NOT NULL DEFAULT 0,
    mae_pips REAL NOT NULL DEFAULT 0, realized_r REAL, realized_r_net REAL,
    opened_at TEXT NOT NULL, partial_at TEXT, closed_at TEXT,
    last_tick_at TEXT NOT NULL, extra_json TEXT
);
"""


@pytest.fixture
def dashboard_db(tmp_path, monkeypatch):
    path = tmp_path / "tracker.db"
    with sqlite3.connect(path) as connection:
        connection.executescript(SCHEMA)
    monkeypatch.setattr(settings, "TRACKER_DB_PATH", str(path))
    return path


@pytest.fixture
def client(dashboard_db):
    return TestClient(app)


def _insert(path, *, setup="london_judas", status="open", realized_r=None, index=1):
    opened_at = f"2026-06-{(index % 28) + 1:02d}T08:{index % 60:02d}:00+00:00"
    with sqlite3.connect(path) as connection:
        cursor = connection.execute(
            """
            INSERT INTO signal_lifecycle (
                setup, direction, pattern, killzone, killzone_match,
                entry, entry_fill, spread_pips, sl, tp1, tp_final, risk_pips, planned_rr,
                status, mfe_pips, mae_pips, realized_r, realized_r_net,
                opened_at, partial_at, closed_at, last_tick_at, extra_json
            ) VALUES (?, 'long', 'test_pattern', 'LONDON', 1,
                      1.0852, 1.0853, 1, 1.0840, 1.0864, 1.0876, 12, 2,
                      ?, 8.2, -4.1, ?, ?, ?, NULL, ?, ?, ?)
            """,
            (
                setup, status, realized_r, realized_r, opened_at,
                opened_at if status.startswith("closed_") else None,
                opened_at, json.dumps({"fixture": index}),
            ),
        )
        return cursor.lastrowid


def test_summary_empty_db(client):
    response = client.get("/api/dashboard/summary")
    assert response.status_code == 200
    assert response.json() == {
        "total_signals": 0, "open": 0, "partial": 0,
        "closed_tp": 0, "closed_be": 0, "closed_sl": 0, "closed_timeout": 0,
        "win_rate": 0.0, "net_r": 0.0, "avg_r": 0.0,
        "first_signal_at": None, "last_signal_at": None,
    }


def test_summary_with_mix(client, dashboard_db):
    for index, (status, realized_r) in enumerate((
        ("open", None), ("closed_tp", 1.5), ("closed_tp", 1.5),
        ("closed_be", 0.5), ("closed_sl", -1.0),
    ), start=1):
        _insert(dashboard_db, status=status, realized_r=realized_r, index=index)
    data = client.get("/api/dashboard/summary").json()
    assert data["total_signals"] == 5
    assert data["open"] == 1
    assert data["closed_tp"] == 2
    assert data["win_rate"] == 0.5
    assert data["net_r"] == pytest.approx(2.5)
    assert data["avg_r"] == pytest.approx(0.625)


def test_summary_counts_timeout_as_closed(client, dashboard_db):
    _insert(dashboard_db, status="closed_timeout", realized_r=0.25)
    data = client.get("/api/dashboard/summary").json()
    assert data["closed_timeout"] == 1
    assert data["avg_r"] == pytest.approx(0.25)


def test_by_setup_aggregates_per_setup(client, dashboard_db):
    _insert(dashboard_db, setup="london_judas", status="closed_tp", realized_r=1.5)
    _insert(dashboard_db, setup="breaker_flip", status="closed_sl", realized_r=-1.0, index=2)
    rows = client.get("/api/dashboard/setups").json()
    assert [row["setup"] for row in rows] == ["breaker_flip", "london_judas"]
    assert all(row["total"] == 1 for row in rows)
    assert rows[1]["win_rate"] == 1.0


def test_list_signals_pagination(client, dashboard_db):
    for index in range(1, 61):
        _insert(dashboard_db, index=index)
    data = client.get("/api/dashboard/signals?limit=20&offset=20").json()
    assert data["total"] == 60
    assert data["limit"] == 20
    assert data["offset"] == 20
    assert len(data["items"]) == 20


def test_list_signals_filter_by_status(client, dashboard_db):
    _insert(dashboard_db, status="open")
    _insert(dashboard_db, status="closed_tp", realized_r=1.5, index=2)
    data = client.get("/api/dashboard/signals?status=open").json()
    assert data["total"] == 1
    assert {item["status"] for item in data["items"]} == {"open"}


def test_list_signals_filter_by_setup(client, dashboard_db):
    _insert(dashboard_db, setup="london_judas")
    _insert(dashboard_db, setup="breaker_flip", index=2)
    data = client.get("/api/dashboard/signals?setup=london_judas").json()
    assert data["total"] == 1
    assert data["items"][0]["setup"] == "london_judas"


def test_signal_detail_returns_full_row(client, dashboard_db):
    signal_id = _insert(dashboard_db, index=7)
    response = client.get(f"/api/dashboard/signals/{signal_id}")
    assert response.status_code == 200
    data = response.json()
    assert set(data) == {
        "id", "setup", "direction", "pattern", "killzone", "killzone_match",
        "entry", "entry_fill", "spread_pips", "sl", "tp1", "tp_final",
        "risk_pips", "planned_rr", "status", "mfe_pips", "mae_pips",
        "realized_r", "realized_r_net", "opened_at", "partial_at",
        "closed_at", "last_tick_at", "extra_json",
    }
    assert data["extra_json"] == {"fixture": 7}


def test_signal_detail_404(client):
    response = client.get("/api/dashboard/signals/999")
    assert response.status_code == 404
    assert response.json()["detail"] == "Signal not found"


def test_export_csv_headers_and_rows(client, dashboard_db):
    _insert(dashboard_db)
    _insert(dashboard_db, index=2)
    response = client.get("/api/dashboard/export.csv")
    lines = response.text.strip().splitlines()
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert lines[0].split(",") == [
        "id", "setup", "direction", "pattern", "killzone", "killzone_match",
        "entry", "entry_fill", "spread_pips", "sl", "tp1", "tp_final",
        "risk_pips", "planned_rr", "status", "mfe_pips", "mae_pips",
        "realized_r", "realized_r_net", "opened_at", "partial_at", "closed_at",
    ]
    assert len(lines) == 3


def test_export_xlsx_returns_valid_workbook(client, dashboard_db):
    _insert(dashboard_db)
    response = client.get("/api/dashboard/export.xlsx")
    workbook = load_workbook(BytesIO(response.content))
    assert response.status_code == 200
    assert workbook["Signals"].max_row == 2
    assert workbook["Signals"]["A1"].value == "id"


def test_dashboard_page_served(client):
    response = client.get("/dashboard")
    assert response.status_code == 200
    assert "EuroTruck Dashboard" in response.text


def test_lifecycle_connection_is_read_only(dashboard_db):
    with ro_connect() as connection:
        with pytest.raises(sqlite3.OperationalError, match="readonly"):
            connection.execute("CREATE TABLE forbidden (id INTEGER)")


def test_missing_tracker_db_returns_503(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "TRACKER_DB_PATH", str(tmp_path / "missing.db"))
    response = TestClient(app).get("/api/dashboard/summary")
    assert response.status_code == 503
    assert response.json()["detail"] == "Tracker DB not yet initialized"
